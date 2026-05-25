import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Учет"

# --- 1. НАСТРОЙКИ (A1:B5) ---
settings = [
    ("Месячный оклад:", 50000),
    ("Часов в рабочем дне:", 8),
    ("Среднее рабочих дней в мес.:", 21.75),
    ("Стоимость 1 часа:", "=B1/(B2*B3)"),
    ("Стоимость 1 рабочего дня:", "=B1/B3")
]
for i, (label, val) in enumerate(settings, start=1):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=val)
    c.number_format = '#,##0.00'

# --- 2. ЗАГОЛОВКИ (Общие для строки 7) ---
# Ежедневный учёт: A7:D7
daily_headers = ["Дата", "День недели", "Часы", "Заработок за день"]
for col, h in enumerate(daily_headers, 1):
    c = ws.cell(row=7, column=col, value=h)
    c.font = Font(bold=True)
    c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    c.alignment = Alignment(horizontal='center')

# Итоговая таблица: G7:L7
summary_headers = ["Месяц", "Всего часов", "Рабочих дней", "ЗП без премии", "Премия (вручную)", "Итого с премией"]
for col, h in enumerate(summary_headers, start=7):
    c = ws.cell(row=7, column=col, value=h)
    c.font = Font(bold=True)
    c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    c.alignment = Alignment(horizontal='center')

# --- 3. ГЕНЕРАЦИЯ ДНЕЙ (A8:D...) ---
RU_DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
month_fills = [
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # Жёлтый
    PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"), # Серый
]

start_date = date(2026, 1, 1)
current_row = 8
current_month = start_date.month
max_data_row = 7

for i in range(365):
    d = start_date + timedelta(days=i)
    
    # Разрыв при смене месяца
    if d.month != current_month:
        current_row += 1
        current_month = d.month

    fill = month_fills[(current_month - 1) % 2]

    # Дата как объект Excel (не текст!)
    cell_date = ws.cell(row=current_row, column=1, value=d)
    cell_date.number_format = 'DD.MM.YYYY'
    
    ws.cell(row=current_row, column=2, value=RU_DAYS[d.weekday()])
    ws.cell(row=current_row, column=3).number_format = '0.00'
    ws.cell(row=current_row, column=4, value=f'=IF(C{current_row}>0, C{current_row}*$B$4, 0)')
    ws.cell(row=current_row, column=4).number_format = '#,##0.00'

    # Закраска строки
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    current_row += 1

max_data_row = current_row - 1

# --- 4. ИТОГИ ПО МЕСЯЦАМ (G8:L...) ---
months = ["Январь","Февраль","Март","Апрель","Май","Июнь",
          "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

summary_start_row = 8
# Диапазоны для формул (динамические)
rng_A = f"$A$8:$A${max_data_row}"
rng_C = f"$C$8:$C${max_data_row}"

for m_idx, name in enumerate(months):
    row = summary_start_row + m_idx * 2  # 8, 10, 12...
    
    # Цвет строки
    fill = month_fills[m_idx % 2]
    for col in range(7, 13):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    ws.cell(row=row, column=7, value=name)
    
    # Часы и дни (с фильтром ISNUMBER, чтобы не считать пустые строки-разрывы)
    ws.cell(row=row, column=8, value=f'=SUMPRODUCT((MONTH({rng_A})={m_idx+1})*(ISNUMBER({rng_A}))*{rng_C})')
    ws.cell(row=row, column=9, value=f'=SUMPRODUCT((MONTH({rng_A})={m_idx+1})*(ISNUMBER({rng_A}))*({rng_C}>0))')
    
    # ЗП без премии = Часы * Стоимость часа
    ws.cell(row=row, column=10, value=f'=H{row}*$B$4')
    
    # Премия (пустая для ручного ввода)
    ws.cell(row=row, column=11, value="")
    
    # Итого с премией
    ws.cell(row=row, column=12, value=f'=J{row}+K{row}')

    for col in (8, 9, 10, 12):
        ws.cell(row=row, column=col).number_format = '#,##0.00'

# --- 5. ГОДОВОЙ ИТОГ ---
total_row = summary_start_row + 12 * 2  # Строка 32
total_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

ws.cell(row=total_row, column=7, value="ИТОГО ЗА ГОД").font = Font(bold=True)
ws.cell(row=total_row, column=7).fill = total_fill

# Диапазоны для сумм (пустые строки игнорируются функцией SUM)
base_rng = f"G{summary_start_row}:G{total_row-2}"
cols_to_sum = [
    (8, f"H{summary_start_row}:H{total_row-2}"), # Часы
    (9, f"I{summary_start_row}:I{total_row-2}"), # Дни
    (10, f"J{summary_start_row}:J{total_row-2}"),# ЗП без премии
    (11, f"K{summary_start_row}:K{total_row-2}"),# Премия
    (12, f"L{summary_start_row}:L{total_row-2}"),# Итого
]

for col, rng in cols_to_sum:
    c = ws.cell(row=total_row, column=col, value=f'=SUM({rng})')
    c.number_format = '#,##0.00'
    c.font = Font(bold=True)
    c.fill = total_fill
    c.alignment = Alignment(horizontal='center')

# --- 6. ШИРИНА СТОЛБЦОВ ---
for col, w in [(1,13), (2,15), (3,10), (4,15), (7,14), (8,12), (9,13), (10,15), (11,14), (12,16)]:
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

wb.save("Табель_учета_2026.xlsx")
print("✅ Готово! Итоговая таблица с G7, премия и расчёты добавлены.")
